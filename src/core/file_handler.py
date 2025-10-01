# -*- coding: utf-8 -*-
"""
文件处理器
负责Excel文件的保存、输出路径管理和文件操作
"""

import pandas as pd
import os
import subprocess
import platform
from typing import Optional, Dict
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Border, Side
from src.utils.logger import get_logger

logger = get_logger("FileHandler")

class FileHandler:
    """
    文件处理器类
    提供Excel文件保存、路径管理、文件操作等功能
    """
    
    def __init__(self):
        self.original_file_path = None
        self.output_directory = None
        
    def set_original_file(self, file_path: str):
        """
        设置原始文件路径
        
        Args:
            file_path: 原始文件路径
        """
        self.original_file_path = file_path
        self.output_directory = os.path.dirname(file_path)
        logger.info(f"设置原始文件: {file_path}")
    
    def set_output_directory(self, directory: str):
        """
        设置输出目录
        
        Args:
            directory: 输出目录路径
        """
        if os.path.exists(directory) and os.path.isdir(directory):
            self.output_directory = directory
            logger.info(f"设置输出目录: {directory}")
        else:
            logger.error(f"输出目录不存在: {directory}")
    
    def generate_output_filename(self, original_path: str, suffix: str = "_processed") -> str:
        """
        生成输出文件名
        
        Args:
            original_path: 原始文件路径
            suffix: 文件名后缀
            
        Returns:
            str: 输出文件路径
        """
        try:
            directory = os.path.dirname(original_path)
            filename = os.path.basename(original_path)
            name, ext = os.path.splitext(filename)
            
            # 生成新文件名
            new_filename = f"{name}{suffix}{ext}"
            output_path = os.path.join(directory, new_filename)
            
            # 如果文件已存在，添加时间戳
            if os.path.exists(output_path):
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                new_filename = f"{name}{suffix}_{timestamp}{ext}"
                output_path = os.path.join(directory, new_filename)
            
            logger.info(f"生成输出文件名: {output_path}")
            return output_path
            
        except Exception as e:
            logger.error(f"生成输出文件名失败: {e}")
            return original_path
    
    def save_to_excel(self, data: pd.DataFrame, output_path: str, sheet_name: str = "Sheet1") -> bool:
        """
        保存数据到Excel文件
        
        Args:
            data: 要保存的数据
            output_path: 输出文件路径
            sheet_name: 工作表名称
            
        Returns:
            bool: 保存是否成功
        """
        try:
            if data is None or data.empty:
                logger.error("没有数据可保存")
                return False
            
            # 确保输出目录存在
            output_dir = os.path.dirname(output_path)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
                logger.info(f"创建输出目录: {output_dir}")
            
            # 保存到Excel文件
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.info(f"数据保存成功: {output_path}")
            logger.info(f"保存数据: {len(data)} 行 x {len(data.columns)} 列")
            return True
            
        except Exception as e:
            logger.error(f"保存Excel文件失败: {e}")
            return False
    
    def save_to_csv(self, data: pd.DataFrame, output_path: str, encoding: str = 'utf-8-sig') -> bool:
        """
        保存数据到CSV文件
        
        Args:
            data: 要保存的数据
            output_path: 输出文件路径
            encoding: 文件编码
            
        Returns:
            bool: 保存是否成功
        """
        try:
            if data is None or data.empty:
                logger.error("没有数据可保存")
                return False
            
            # 确保输出目录存在
            output_dir = os.path.dirname(output_path)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
                logger.info(f"创建输出目录: {output_dir}")
            
            # 保存到CSV文件
            data.to_csv(output_path, index=False, encoding=encoding)
            
            logger.info(f"CSV文件保存成功: {output_path}")
            return True
            
        except Exception as e:
            logger.error(f"保存CSV文件失败: {e}")
            return False
    
    def save_excel_with_format(self, source_path: str, target_path: str, 
                              data: pd.DataFrame, sheet_name: str) -> bool:
        """
        保存Excel文件并保留原始格式（列宽、行高、样式等）
        
        Args:
            source_path: 源文件路径
            target_path: 目标文件路径
            data: 修改后的数据
            sheet_name: 要修改的工作表名称
            
        Returns:
            bool: 操作是否成功
        """
        try:
            # 加载原始工作簿
            source_wb = load_workbook(source_path)
            
            # 检查目标工作表是否存在
            if sheet_name in source_wb.sheetnames:
                target_ws = source_wb[sheet_name]
                
                # 保存原始的列宽和行高信息
                original_column_widths = {}
                original_row_heights = {}
                
                # 获取原始列宽
                for col_letter, col_dimension in target_ws.column_dimensions.items():
                    if col_dimension.width:
                        original_column_widths[col_letter] = col_dimension.width
                
                # 获取原始行高
                for row_num, row_dimension in target_ws.row_dimensions.items():
                    if row_dimension.height:
                        original_row_heights[row_num] = row_dimension.height
                
                # 清除工作表内容但保留格式
                target_ws.delete_rows(1, target_ws.max_row)
                
                # 写入新数据
                for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        target_ws.cell(row=r_idx, column=c_idx, value=value)
                
                # 恢复列宽
                for col_letter, width in original_column_widths.items():
                    target_ws.column_dimensions[col_letter].width = width
                
                # 恢复行高
                for row_num, height in original_row_heights.items():
                    if row_num <= target_ws.max_row:
                        target_ws.row_dimensions[row_num].height = height
                
                # 保存工作簿
                source_wb.save(target_path)
                source_wb.close()
                
                logger.info(f"Excel文件保存成功（保留格式）: {target_path}")
                return True
            else:
                logger.error(f"工作表 {sheet_name} 不存在于源文件中")
                return False
                
        except Exception as e:
            logger.error(f"保存Excel文件（保留格式）失败: {e}")
            return False
    
    def save_excel_with_format_and_border(self, source_path: str, target_path: str, 
                                         data: pd.DataFrame, sheet_name: str, 
                                         add_border: bool = True) -> bool:
        """
        保存Excel文件并保留原始格式，可选择添加边框
        
        Args:
            source_path: 源文件路径
            target_path: 目标文件路径
            data: 修改后的数据
            sheet_name: 要修改的工作表名称
            add_border: 是否添加边框
            
        Returns:
            bool: 操作是否成功
        """
        try:
            # 加载原始工作簿
            source_wb = load_workbook(source_path)
            
            # 检查目标工作表是否存在
            if sheet_name in source_wb.sheetnames:
                target_ws = source_wb[sheet_name]
                
                # 保存原始的列宽和行高信息
                original_column_widths = {}
                original_row_heights = {}
                
                # 获取原始列宽
                for col_letter, col_dimension in target_ws.column_dimensions.items():
                    if col_dimension.width:
                        original_column_widths[col_letter] = col_dimension.width
                
                # 获取原始行高
                for row_num, row_dimension in target_ws.row_dimensions.items():
                    if row_dimension.height:
                        original_row_heights[row_num] = row_dimension.height
                
                # 清除工作表内容但保留格式
                target_ws.delete_rows(1, target_ws.max_row)
                
                # 写入新数据
                for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                    for c_idx, value in enumerate(row, 1):
                        cell = target_ws.cell(row=r_idx, column=c_idx, value=value)
                        
                        # 添加边框
                        if add_border:
                            thin_border = Border(
                                left=Side(style='thin'),
                                right=Side(style='thin'),
                                top=Side(style='thin'),
                                bottom=Side(style='thin')
                            )
                            cell.border = thin_border
                
                # 恢复列宽
                for col_letter, width in original_column_widths.items():
                    target_ws.column_dimensions[col_letter].width = width
                
                # 恢复行高
                for row_num, height in original_row_heights.items():
                    if row_num <= target_ws.max_row:
                        target_ws.row_dimensions[row_num].height = height
                
                # 保存工作簿
                source_wb.save(target_path)
                source_wb.close()
                
                border_info = "（含边框）" if add_border else "（无边框）"
                logger.info(f"Excel文件保存成功{border_info}: {target_path}")
                return True
            else:
                logger.error(f"工作表 {sheet_name} 不存在于源文件中")
                return False
                
        except Exception as e:
            logger.error(f"保存Excel文件（保留格式和边框）失败: {e}")
            return False
    
    def copy_file_with_modifications(self, source_path: str, target_path: str, 
                                   data: pd.DataFrame, sheet_name: str) -> bool:
        """
        复制文件并修改指定工作表（保留原始格式）
        
        Args:
            source_path: 源文件路径
            target_path: 目标文件路径
            data: 修改后的数据
            sheet_name: 要修改的工作表名称
            
        Returns:
            bool: 操作是否成功
        """
        try:
            # 使用新的格式保留方法
            return self.save_excel_with_format(source_path, target_path, data, sheet_name)
            
        except Exception as e:
            logger.error(f"复制并修改文件失败: {e}")
            return False
    
    def open_file(self, file_path: str) -> bool:
        """
        打开文件
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 操作是否成功
        """
        try:
            if not os.path.exists(file_path):
                logger.error(f"文件不存在: {file_path}")
                return False
            
            system = platform.system()
            
            if system == "Windows":
                os.startfile(file_path)
            elif system == "Darwin":  # macOS
                subprocess.run(["open", file_path])
            else:  # Linux
                subprocess.run(["xdg-open", file_path])
            
            logger.info(f"打开文件: {file_path}")
            return True
            
        except Exception as e:
            logger.error(f"打开文件失败: {e}")
            return False
    
    def open_file_location(self, file_path: str) -> bool:
        """
        打开文件所在位置
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 操作是否成功
        """
        try:
            if not os.path.exists(file_path):
                logger.error(f"文件不存在: {file_path}")
                return False
            
            directory = os.path.dirname(file_path)
            system = platform.system()
            
            if system == "Windows":
                subprocess.run(["explorer", "/select,", file_path])
            elif system == "Darwin":  # macOS
                subprocess.run(["open", "-R", file_path])
            else:  # Linux
                subprocess.run(["xdg-open", directory])
            
            logger.info(f"打开文件位置: {directory}")
            return True
            
        except Exception as e:
            logger.error(f"打开文件位置失败: {e}")
            return False
    
    def get_file_size(self, file_path: str) -> Optional[float]:
        """
        获取文件大小（MB）
        
        Args:
            file_path: 文件路径
            
        Returns:
            Optional[float]: 文件大小（MB）
        """
        try:
            if not os.path.exists(file_path):
                return None
                
            size_bytes = os.path.getsize(file_path)
            size_mb = round(size_bytes / (1024 * 1024), 2)
            return size_mb
            
        except Exception as e:
            logger.error(f"获取文件大小失败: {e}")
            return None
    
    def validate_file_path(self, file_path: str) -> bool:
        """
        验证文件路径是否有效
        
        Args:
            file_path: 文件路径
            
        Returns:
            bool: 路径是否有效
        """
        try:
            # 检查路径格式
            if not file_path or not isinstance(file_path, str):
                return False
            
            # 检查目录是否存在
            directory = os.path.dirname(file_path)
            if not os.path.exists(directory):
                return False
            
            # 检查是否有写入权限
            if not os.access(directory, os.W_OK):
                return False
            
            return True
            
        except Exception as e:
            logger.error(f"验证文件路径失败: {e}")
            return False
    
    def create_backup(self, file_path: str) -> Optional[str]:
        """
        创建文件备份
        
        Args:
            file_path: 原文件路径
            
        Returns:
            Optional[str]: 备份文件路径
        """
        try:
            if not os.path.exists(file_path):
                logger.error(f"原文件不存在: {file_path}")
                return None
            
            # 生成备份文件名
            directory = os.path.dirname(file_path)
            filename = os.path.basename(file_path)
            name, ext = os.path.splitext(filename)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            backup_filename = f"{name}_backup_{timestamp}{ext}"
            backup_path = os.path.join(directory, backup_filename)
            
            # 复制文件
            import shutil
            shutil.copy2(file_path, backup_path)
            
            logger.info(f"创建备份文件: {backup_path}")
            return backup_path
            
        except Exception as e:
            logger.error(f"创建备份文件失败: {e}")
            return None
    
    def save_multiple_sheets_to_excel(self, sheet_data: Dict[str, pd.DataFrame], 
                                    output_path: str, preserve_format: bool = True,
                                    source_path: Optional[str] = None) -> bool:
        """
        保存多个工作表到Excel文件
        
        Args:
            sheet_data: 包含所有工作表数据的字典，键为工作表名，值为DataFrame
            output_path: 输出文件路径
            preserve_format: 是否保留原始格式
            source_path: 源文件路径（用于保留格式）
            
        Returns:
            bool: 保存是否成功
        """
        try:
            if not sheet_data:
                logger.error("没有工作表数据可保存")
                return False
            
            # 确保输出目录存在
            output_dir = os.path.dirname(output_path)
            if not os.path.exists(output_dir):
                os.makedirs(output_dir)
                logger.info(f"创建输出目录: {output_dir}")
            
            if preserve_format and source_path and os.path.exists(source_path):
                # 保留原始格式的保存方式
                return self._save_with_format_preservation(sheet_data, output_path, source_path)
            else:
                # 普通保存方式
                return self._save_without_format_preservation(sheet_data, output_path)
                
        except Exception as e:
            logger.error(f"保存多工作表Excel文件失败: {e}")
            return False
    
    def _save_with_format_preservation(self, sheet_data: Dict[str, pd.DataFrame], 
                                     output_path: str, source_path: str) -> bool:
        """
        保存Excel文件并保留原始格式
        
        Args:
            sheet_data: 工作表数据字典
            output_path: 输出文件路径
            source_path: 源文件路径
            
        Returns:
            bool: 保存是否成功
        """
        try:
            # 加载原始工作簿
            source_wb = load_workbook(source_path)
            
            # 更新每个工作表的数据
            for sheet_name, data in sheet_data.items():
                if sheet_name in source_wb.sheetnames:
                    target_ws = source_wb[sheet_name]
                    
                    # 保存原始的列宽和行高信息
                    original_column_widths = {}
                    original_row_heights = {}
                    
                    # 获取原始列宽
                    for col_letter, col_dimension in target_ws.column_dimensions.items():
                        if col_dimension.width:
                            original_column_widths[col_letter] = col_dimension.width
                    
                    # 获取原始行高
                    for row_num, row_dimension in target_ws.row_dimensions.items():
                        if row_dimension.height:
                            original_row_heights[row_num] = row_dimension.height
                    
                    # 清除工作表内容但保留格式
                    target_ws.delete_rows(1, target_ws.max_row)
                    
                    # 写入新数据
                    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            target_ws.cell(row=r_idx, column=c_idx, value=value)
                    
                    # 恢复列宽
                    for col_letter, width in original_column_widths.items():
                        target_ws.column_dimensions[col_letter].width = width
                    
                    # 恢复行高
                    for row_num, height in original_row_heights.items():
                        if row_num <= target_ws.max_row:
                            target_ws.row_dimensions[row_num].height = height
                else:
                    # 如果工作表不存在，创建新的工作表
                    new_ws = source_wb.create_sheet(sheet_name)
                    for r_idx, row in enumerate(dataframe_to_rows(data, index=False, header=True), 1):
                        for c_idx, value in enumerate(row, 1):
                            new_ws.cell(row=r_idx, column=c_idx, value=value)
            
            # 保存工作簿
            source_wb.save(output_path)
            source_wb.close()
            
            logger.info(f"Excel文件保存成功（保留格式）: {output_path}")
            logger.info(f"保存工作表数量: {len(sheet_data)}")
            return True
            
        except Exception as e:
            logger.error(f"保存Excel文件（保留格式）失败: {e}")
            return False
    
    def _save_without_format_preservation(self, sheet_data: Dict[str, pd.DataFrame], 
                                        output_path: str) -> bool:
        """
        保存Excel文件不保留原始格式
        
        Args:
            sheet_data: 工作表数据字典
            output_path: 输出文件路径
            
        Returns:
            bool: 保存是否成功
        """
        try:
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                for sheet_name, data in sheet_data.items():
                    if data is not None and not data.empty:
                        data.to_excel(writer, sheet_name=sheet_name, index=False)
            
            logger.info(f"Excel文件保存成功: {output_path}")
            logger.info(f"保存工作表数量: {len(sheet_data)}")
            return True
            
        except Exception as e:
            logger.error(f"保存Excel文件失败: {e}")
            return False